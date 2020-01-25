import requests
import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

SPREADSHEET_NAME = 'test.xlsx'


if __name__ == '__main__':
    # Fetching currencies
    soup = BeautifulSoup(requests.get('https://www.tgju.org/chart/price_dollar_rl').text, "html.parser")
    dollar_price = soup.select_one(
        '#main > div.profile-container.container > div:nth-child(2) > ul > li:nth-child(1) > span'
    ).text

    soup = BeautifulSoup(requests.get('http://www.tgju.org/chart/price_eur').text, "html.parser")
    euro_price = soup.select_one(
        '#main > div.profile-container.container > div:nth-child(2) > ul > li:nth-child(1) > span'
    ).text

    soup = BeautifulSoup(requests.get('http://www.tgju.org/chart/price_aed').text, "html.parser")
    dirham_price = soup.select_one(
        '#main > div.profile-container.container > div:nth-child(2) > ul > li:nth-child(1) > span'
    ).text

    soup = BeautifulSoup(requests.get('http://www.tgju.org/chart/price_cny').text, "html.parser")
    yuan_price = soup.select_one(
        '#main > div.profile-container.container > div:nth-child(2) > ul > li:nth-child(1) > span'
    ).text

    # Fetching petrol prices
    soup = BeautifulSoup(requests.get('http://www.tgju.org/chart/energy-crude-oil').text, "html.parser")
    crude_price = soup.select_one(
        '#main > div.profile-container.container > div:nth-child(2) > ul > li:nth-child(1) > span'
    ).text

    soup = BeautifulSoup(requests.get('http://www.tgju.org/chart/energy-brent-oil').text, "html.parser")
    brent_price = soup.select_one(
        '#main > div.profile-container.container > div:nth-child(2) > ul > li:nth-child(1) > span'
    ).text

    soup = BeautifulSoup(requests.get('http://www.tgju.org/chart/oil_opec').text, "html.parser")
    opec_price = soup.select_one(
        '#main > div.profile-container.container > div:nth-child(2) > ul > li:nth-child(1) > span'
    ).text

    soup = BeautifulSoup(requests.get('http://www.tgju.org/chart/energy-heating-oil').text, "html.parser")
    mazut_price = soup.select_one(
        '#main > div.profile-container.container > div:nth-child(2) > ul > li:nth-child(1) > span'
    ).text

    # Fetching global gold prices
    soup = BeautifulSoup(requests.get('http://www.tgju.org/chart/ons').text, "html.parser")
    gold_dollars = soup.select_one(
        '#main > div.profile-container.container > div:nth-child(2) > ul > li:nth-child(1) > span'
    ).text

    # Writing the scraped data into a spreadsheet
    scraped_data = [datetime.datetime.now().date(), dollar_price, euro_price, dirham_price, yuan_price, crude_price, brent_price, opec_price, mazut_price, gold_dollars]

    # Checking if the spreadsheet exists, if so, append the scraped data to it
    try:
        wb = load_workbook(SPREADSHEET_NAME)
        ws = wb.active

        ws.insert_rows(2)
        for i in range(len(scraped_data)):
            ws.cell(row=2, column=i+1).value = scraped_data[i]

        wb.save(SPREADSHEET_NAME)
        print('Data added to ' + SPREADSHEET_NAME)
    except FileNotFoundError:   # Spreadsheet does not exist in the root folder, creating a new one and appending to it
        wb = Workbook()
        ws = wb.active
        ws.append(['Datetime', 'Dollar Price', 'Euro Price', 'Dirham Price', 'Yuan Price', 'Crude Price', 'Brent Price', 'Opec Price', 'Mazut Price', 'Global Gold Price ($)'])
        ws.append(scraped_data)
        wb.save(SPREADSHEET_NAME)
        print('Data written in ' + SPREADSHEET_NAME)
